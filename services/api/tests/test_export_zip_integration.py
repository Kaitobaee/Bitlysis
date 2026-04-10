"""Phase 8 — mở ZIP, đếm file/sheet; heavy export cần export/start."""

from __future__ import annotations

import json
import uuid
import zipfile
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.config import Settings, get_settings
from app.main import app
from app.services.provenance import build_run_manifest, write_manifest


def _seed_job(tmp_path: Path) -> str:
    jid = str(uuid.uuid4())
    csv_name = f"{jid}.csv"
    (tmp_path / csv_name).write_text("x,y,z\n1,10,aa\n2,20,bb\n", encoding="utf-8")
    man_rel = f"{jid}.manifest.json"
    man = build_run_manifest(jid, 1)
    write_manifest(tmp_path / man_rel, man)
    meta = {
        "job_id": jid,
        "status": "succeeded",
        "stored_as": csv_name,
        "original_filename": "t.csv",
        "columns": ["x", "y", "z"],
        "row_preview_count": 2,
        "size_bytes": 40,
        "uploaded_at": "2024-01-01T00:00:00+00:00",
        "profiling_engine_version": 1,
        "manifest_stored_as": man_rel,
        "result_summary": {"engine": "python_stats", "hypothesis_table": []},
    }
    (tmp_path / f"{jid}.meta.json").write_text(
        json.dumps(meta, ensure_ascii=False),
        encoding="utf-8",
    )
    return jid


@pytest.fixture
def export_client(tmp_path: Path):
    jid = _seed_job(tmp_path)
    settings = Settings(
        upload_dir=tmp_path,
        export_zip_heavy_threshold_bytes=50_000_000,
        export_include_plotly=False,
    )
    app.dependency_overrides[get_settings] = lambda: settings
    yield TestClient(app), jid, tmp_path
    app.dependency_overrides.clear()


def test_matplotlib_chart_preview_endpoint_returns_png(export_client):
    client, jid, _tmp = export_client
    r = client.get(f"/v1/jobs/{jid}/charts/matplotlib")
    assert r.status_code == 200, r.text
    assert r.headers.get("content-type", "").startswith("image/png")
    assert len(r.content) > 500


def test_export_zip_contains_docs_layout_and_excel_sheets(export_client):
    client, jid, _tmp = export_client
    r = client.post(f"/v1/jobs/{jid}/export")
    assert r.status_code == 200, r.text
    zf = zipfile.ZipFile(BytesIO(r.content))
    names = set(zf.namelist())
    assert "run_manifest.json" in names
    assert "docs/charts/matplotlib_series.png" in names
    assert "docs/tables/summary_tables.pdf" in names
    assert "docs/report.docx" in names
    assert "docs/data/workbook.xlsx" in names
    man = json.loads(zf.read("run_manifest.json").decode("utf-8"))
    assert man.get("export", {}).get("phase") == 8
    wb = load_workbook(BytesIO(zf.read("docs/data/workbook.xlsx")))
    assert set(wb.sheetnames) == {"data_clean", "results_raw"}
    assert len(wb.sheetnames) == 2
    ws = wb["data_clean"]
    assert ws.max_row >= 2
    ws2 = wb["results_raw"]
    assert ws2.max_row >= 2


def test_heavy_export_requires_start_then_succeeds(tmp_path: Path) -> None:
    jid = _seed_job(tmp_path)
    settings = Settings(
        upload_dir=tmp_path,
        export_zip_heavy_threshold_bytes=1,
        export_include_plotly=False,
    )
    app.dependency_overrides[get_settings] = lambda: settings
    try:
        with TestClient(app) as client:
            r0 = client.post(f"/v1/jobs/{jid}/export")
            assert r0.status_code == 409
            det = r0.json().get("details")
            assert isinstance(det, dict)
            assert det.get("code") == "heavy_export_requires_export_phase"
            r1 = client.post(f"/v1/jobs/{jid}/export/start")
            assert r1.status_code == 202
            r2 = client.post(f"/v1/jobs/{jid}/export")
            assert r2.status_code == 200
            zf = zipfile.ZipFile(BytesIO(r2.content))
            assert "run_manifest.json" in zf.namelist()
    finally:
        app.dependency_overrides.clear()


def test_export_start_requires_succeeded(tmp_path: Path) -> None:
    jid = str(uuid.uuid4())
    (tmp_path / f"{jid}.csv").write_text("a\n1\n", encoding="utf-8")
    meta = {
        "job_id": jid,
        "status": "uploaded",
        "stored_as": f"{jid}.csv",
        "original_filename": "u.csv",
        "columns": ["a"],
        "row_preview_count": 1,
        "size_bytes": 5,
        "uploaded_at": "2024-01-01T00:00:00+00:00",
    }
    (tmp_path / f"{jid}.meta.json").write_text(json.dumps(meta), encoding="utf-8")
    settings = Settings(upload_dir=tmp_path)
    app.dependency_overrides[get_settings] = lambda: settings
    try:
        with TestClient(app) as client:
            r = client.post(f"/v1/jobs/{jid}/export/start")
        assert r.status_code == 409
    finally:
        app.dependency_overrides.clear()


def test_export_download_after_build(export_client) -> None:
    client, jid, _tmp = export_client
    r1 = client.post(f"/v1/jobs/{jid}/export")
    assert r1.status_code == 200
    r2 = client.get(f"/v1/jobs/{jid}/export/download")
    assert r2.status_code == 200
    assert r2.content == r1.content
