"""Tạo file .xlsx fixtures cho tests (chạy từ thư mục services/api)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

FIXTURES = Path(__file__).resolve().parent.parent / "tests" / "fixtures"


def main() -> None:
    FIXTURES.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws.append(["id", "product", "qty", "price"])
    ws.append([1, "A", 2, 10.5])
    ws.append([2, "B", 1, 20.0])
    ws.append([3, "A", 5, 10.5])
    wb.save(FIXTURES / "sample_clean.xlsx")

    wb2 = Workbook()
    s1 = wb2.active
    assert s1 is not None
    s1.title = "Sales"
    s1.append(["date", "region", "amount"])
    s1.append(["2024-01-01", "North", 100])
    s1.append(["2024-01-02", "South", 150])
    s2 = wb2.create_sheet("Summary")
    s2["A1"] = "Quarter"
    s2["B1"] = "Total"
    s2.merge_cells("A2:B2")
    s2["A2"] = "Q1 aggregate (merged header cell)"
    wb2.save(FIXTURES / "sample_multi_sheet_merged.xlsx")

    print(f"Wrote fixtures under {FIXTURES}")


if __name__ == "__main__":
    main()
